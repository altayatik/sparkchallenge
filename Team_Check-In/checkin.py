import openpyxl as xl
import time
import os
from os import system

path = "./Registration Data.xlsx"
PUID_row = "A"
name_row = "B"
course_row = "C"
team_row = "D"
table_row = "E"
date_row = "F"
check_in_row = "G"
check_out_row = "H"
difference_in_time_row = "I"
# sheet = wb["Roster (4)"]
#path = "Team_list.xlsx"

#given the ID, find the row number in column A, and return the respective Name, Team, and Table number in the same row in columns B, C, and D
#if Team and Table number are empty, return "N/A"

def find_row(id):
    wb = xl.load_workbook(path)
    sheet = wb["Roster (4)"]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 1)
        if str(cell.value) == str(id):
            return sheet.cell(i, 2).value, sheet.cell(i, 3).value, sheet.cell(i, 4).value, sheet.cell(i, 5).value
    return "Not Found", "Not Found", "Not Found", "Not Found"

#given the ID, find the row number in column A, and update the same row in column F with the current time. If the student has already checked in, 
#update row G with the current time and calculate the time difference between the two times. If the student has already checked out, return "Already Checked Out"
def check_in(id):
    wb = xl.load_workbook(path)
    sheet = wb["Roster (4)"]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 1)
        if str(cell.value) == id:
            if (sheet.cell(i, 7).value == None) or (sheet.cell(i, 7).value == ""):
                sheet.cell(i, 7).value = "Checked In"
                wb.save(path)
                return "Checked In"
            elif sheet.cell(i, 8).value == None or sheet.cell(i, 8).value == "":
                sheet.cell(i, 8).value = "Checked Out"
                wb.save(path)
                return "Checked Out"
            else:
                return "Already Checked Out"
    return "Not Found"

#students who are here for extra credit do not have a team or table number, instead they have a class in Column E and a check in time in Column F and a check out time in Column G
#if the student has no check in time, update the check in time in Column F with the current time. If the student has a check in time but no check out time, update the check out time in Column G with the current time and calculate the time difference between the two times. If the student has already checked out, return "Already Checked Out"
def extra_credit(id):
    wb = xl.load_workbook(path)
    sheet = wb["Roster (4)"]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 1)
        if str(cell.value) == id:
            if (sheet.cell(i, 7).value == None) or (sheet.cell(i, 7).value == ""):
                current_time = time.strftime("%H:%M:%S")
                sheet.cell(i, 7).value = current_time
                wb.save(path)
                return "Checked In"
            elif sheet.cell(i, 8).value == None or sheet.cell(i, 8).value == "":
                current_time = time.strftime("%H:%M:%S")
                #calculate time difference
                check_in_time = sheet.cell(i, 7).value
                check_in_time = check_in_time.split(":")
                check_out_time = current_time.split(":")
                check_in_time = int(check_in_time[0]) * 3600 + int(check_in_time[1]) * 60 + int(check_in_time[2])
                check_out_time = int(check_out_time[0]) * 3600 + int(check_out_time[1]) * 60 + int(check_out_time[2])
                time_difference = check_out_time - check_in_time
                time_difference = time.strftime("%H:%M:%S", time.gmtime(time_difference))
                hours = time_difference.split(":")
                hours = int(hours[0])
                minutes = time_difference.split(":")
                minutes = int(minutes[1])
                if(hours == 0 and minutes < 30):
                    return_str = "Not Enough Time. Please stay for at least 30 minutes to be eligible for Extra Credit." + " Time stayed:" + time_difference
                    return return_str

                sheet.cell(i, 8).value = current_time
                sheet.cell(i, 9).value = time_difference
                wb.save(path)
                return "Checked Out"
            else:
                return "Already Checked Out"
    return "Not Found"

def reset_cli(x = 3):
    #clear the screen
    time.sleep(x)
    system('cls' if os.name == 'nt' else 'clear')
    print("Welcome to the Check-In System!")
#Driver function. 
def verify_check_in():
    PUID = ''
    while PUID == '':
        PUID = input('Swipe PUID...')
        PUID = PUID.strip()
        PUID = PUID.split('=')

        try:
            PUID = PUID[2]
        except IndexError:
            print('ERROR READING CARD')
            PUID = '' # Your while loop is running when PUID is empty --> Need to reset to prevent continuing to next part even though the swipe failed because it still produces something
        
        PUID = PUID[1:] #commented out so I don't have to go brain damage mode to input my stuff and test
        #uncomment the following ^ line based on verification file: PUID = PUID[1:] if verification file is in the format:
        # 12312312 as opposed to 0012312312
        if PUID == 'exit':
            print('Exit Success')
            exit(0)
    #clear the screen
    system('cls' if os.name == 'nt' else 'clear')
    name, team, table, course = find_row(PUID)
    if(name == "Not Found"):
        PUID = PUID[1:]
        name, course, team, table = find_row(PUID)
    if(team == None):
        print("Name: " + str(name))
        check_in_status = extra_credit(PUID)
        print("Course: " + str(course))
        print("Check In Status: " + str(check_in_status))
    else:
        check_in_status = check_in(PUID)
        print("Name: " + str(name)) 
        print("Team: " + str(team))
        print("Table: " + str(table))
        print("Check In Status: " + str(check_in_status))
    #print("Resetting...")

    #put a delay here so the user can read the output. Change the number in the function to change the delay
    #reset_cli(5)

    
#initialize and run the driver function in an infinite loop
if __name__ == "__main__":
    print("Welcome to the Check-In Station!")

    # Use '==' if you want to add a student manually
    # Example: ==0012345678
    
    while True:
        verify_check_in()



