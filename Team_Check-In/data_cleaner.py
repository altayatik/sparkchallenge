import openpyxl as xl

def get_29401_attendees():
    wb = xl.load_workbook("./Registration Data.xlsx")
    sheet = wb["Roster (4)"]
    ECE29401 = {}
    #store as dictionary with key = ID, value = [Name]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 3)
        if (str(cell.value) == "ECE 29401" and (sheet.cell(i,8).value != None)):
            ECE29401[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
            #print(sheet.cell(i,8).value)
        #else if check if student was marked as present for 270
        elif (str(cell.value) == "ECE 29401"):
            #search for student in the same row in column A and confirm that column C is ECE 270 and column H is not empty
            for j in range(1, sheet.max_row + 1):
                cell = sheet.cell(j, 1)
                if (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 27000" and (sheet.cell(j,8) != None or sheet.cell(j,8) != "")):
                    ECE29401[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
                #check the same for 362
                elif (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 36200" and (sheet.cell(j,8) != None or sheet.cell(j,8) != "")):
                    ECE29401[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
    return ECE29401
def get_29401_did_not_attend():
    wb = xl.load_workbook("./Registration Data.xlsx")
    sheet = wb["Roster (4)"]
    did_not_attend = {}
    #store as dictionary with key = ID, value = [Name]
    #get all students who are in 29401
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 3)
        if (str(cell.value) == "ECE 29401"):
            did_not_attend[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
    #remove students who were marked as present for 29401
    ECE29401 = get_29401_attendees()
    for key in ECE29401:
        did_not_attend.pop(key)
    return did_not_attend

def get_270_attendees():
    #same as 29401 attendees but for 270
    wb = xl.load_workbook("./Registration Data.xlsx")
    sheet = wb["Roster (4)"]
    ECE270 = {}
    #store as dictionary with key = ID, value = [Name]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 3)
        if (str(cell.value) == "ECE 27000" and (sheet.cell(i,7).value != None)):
            ECE270[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
            #print(sheet.cell(i,8).value)
        #else if check if student was marked as present for 29401
        elif (str(cell.value) == "ECE 27000"):
            #search for student in the same row in column A and confirm that column C is ECE 29401 and column H is not empty
            for j in range(1, sheet.max_row + 1):
                cell = sheet.cell(j, 1)
                if (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 29401" and (sheet.cell(j,7) != None or sheet.cell(j,8) != "")):
                    ECE270[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
                #check the same for 362
                elif (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 36200" and (sheet.cell(j,7) != None or sheet.cell(j,8) != "")):
                    ECE270[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
    return ECE270

def get_362_attendees():
    wb = xl.load_workbook("./Registration Data.xlsx")
    sheet = wb["Roster (4)"]
    ECE362 = {}
    #store as dictionary with key = ID, value = [Name]
    for i in range(1, sheet.max_row + 1):
        cell = sheet.cell(i, 3)
        if (str(cell.value) == "ECE 36200" and (sheet.cell(i,8).value != None)):
            ECE362[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
            #print(sheet.cell(i,8).value)
        #else if check if student was marked as present for 29401
        elif (str(cell.value) == "ECE 36200"):
            #search for student in the same row in column A and confirm that column C is ECE 29401 and column H is not empty
            for j in range(1, sheet.max_row + 1):
                cell = sheet.cell(j, 1)
                if (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 29401" and (sheet.cell(j,8) != None or sheet.cell(j,8) != "")):
                    ECE362[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
                #check the same for 270
                elif (str(cell.value) == str(sheet.cell(i, 1).value) and str(sheet.cell(j, 3).value) == "ECE 27000" and (sheet.cell(j,8) != None or sheet.cell(j,8) != "")):
                    ECE362[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
    return ECE362

def read_270_roster():
    #open roster.txt
    roster = open("roster.txt", "r")
    #read the file and store the lines in a list
    roster_list = roster.readlines()
    #close the file
    roster.close()
    #pop first line
    roster_list.pop(0)

    #list is in the form 033584584 user
    #store as dictionary with key = ID, value = [user]
    ECE270 = {}
    for i in range(len(roster_list)):
        #split the line into ID and Name
        split_line = roster_list[i].split(" ")
        #store ID and Name in dictionary
        ECE270[int(split_line[0])] = split_line[1]
    return ECE270

def get_270_user(PUID):
    roster_270 = read_270_roster()
    #append a 0 to the front of the PUID if it is not 9 digits long
    #PUID = "0" + str(PUID)
    #return the user associated with the PUID
    return roster_270[PUID]

if __name__ == "__main__":
    #make a new CSV called "29401_attendees.csv" and write the dictionary to it
    import csv
    with open('29401_attendees.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        for key, value in get_29401_attendees().items():
            writer.writerow([key, value])
    print("Length of 29401_attendees.csv: ", len(get_29401_attendees()))
    
    #makee a new CSV called "29401_did_not_attend.csv" and write the students who did not attend to it
    with open('29401_did_not_attend.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        for key, value in get_29401_did_not_attend().items():
            writer.writerow([key, value])

    #make a new CSV called "270_attendees.csv" and write the dictionary to it
    with open('270_attendees.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        for key, value in get_270_attendees().items():
            writer.writerow([key, get_270_user(key)])                
    print("Length of 270_attendees.csv: ", len(get_270_attendees()))

    #make a new CSV called "362_attendees.csv" and write the dictionary to it
    with open('362_attendees.csv', 'w') as csv_file:
        writer = csv.writer(csv_file)
        for key, value in get_362_attendees().items():
            writer.writerow([key, value])
    print("Length of 362_attendees.csv: ", len(get_362_attendees()))


